# AUTOMATIZANDO PLANILHAS COM PYTHON

# imports
from openpyxl import Workbook, load_workbook

# Criando um arquivo do tipo .xlsx
arquivo = Workbook()
arquivo.save('planilha.xlsx')

# Lendo um arquivo do tipo .xlsx
arq = load_workbook('data.xlsx')
plan = arq['Sheet']

# Retorna uma tupla de valores que pode ser percorrida com um laço for
print(plan['A'])


# Percorrendo um tupla
for celula in plan['A']:
    print(celula.value)
